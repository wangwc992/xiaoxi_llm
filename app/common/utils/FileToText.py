import tempfile
import requests
import pytesseract
from io import BytesIO
import openpyxl
from PIL import Image, ImageOps, ImageEnhance
import fitz
from docx import Document
import re


class FileToText:
    def __init__(self, tesseract_cmd=None):
        if tesseract_cmd:
            pytesseract.pytesseract.tesseract_cmd = tesseract_cmd

    def download_file(self, pdf_url):
        if pdf_url.startswith("http://") or pdf_url.startswith("https://"):
            response = requests.get(pdf_url)
            response.raise_for_status()
            return BytesIO(response.content)
        else:
            return open(pdf_url, 'rb')

    def get_file_extension(self, file_obj):
        # 尝试从文件名中获取扩展名
        try:
            return file_obj.split('.')[-1].lower()
        except AttributeError:
            return ""

    def fileToString(self, file_obj, file_type=None):
        text = ""
        if file_type == "pdf" or file_type == "pptx":
            return self.pdfToString(file_obj)
        elif file_type == "jpg" or file_type == "png" or file_type == "gif" or file_type == "jpeg":
            return self.imageToString(file_obj)
        elif file_type == "doc" or file_type == "docx":
            return self.docToString(file_obj)
        elif file_type == "xls" or file_type == "xlsx":
            return self.xlsToString(file_obj)
        else:
            return ""

    def docToString(self, file_obj):
        text = ""
        try:
            doc = Document(file_obj)
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
        except Exception as e:
            print("Error:", e)
        return text

    def xlsToString(self, file_obj):
        text = ""
        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    for cell_value in row:
                        if cell_value is not None:
                            text += str(cell_value) + "\n"
        except Exception as e:
            print("Error:", e)
        return text

    def imageToString(self, file_obj):
        text = ""
        img = Image.open(file_obj)
        img = ImageOps.grayscale(img)
        enhancer = ImageEnhance.Contrast(img)
        img = enhancer.enhance(5)
        img = img.resize((img.width * 2, img.height * 2), Image.LANCZOS)
        enhancer = ImageEnhance.Sharpness(img)
        img = enhancer.enhance(2)
        img = ImageOps.invert(img)
        img = img.point(lambda x: 0 if x < 128 else 255)
        text += pytesseract.image_to_string(img, lang='chi_sim+eng')
        return text

    def pdfToString(self, file_obj):
        with tempfile.NamedTemporaryFile(delete=False) as tmp_pdf_file:
            tmp_pdf_file.write(file_obj.read())
        tmp_pdf_file_path = tmp_pdf_file.name
        doc = fitz.open(tmp_pdf_file_path)
        text = ""
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            page_text = page.get_text()
            if page_text.strip():
                text += page_text
            else:
                # 图片型，转为imageToString支持的形式
                page_image = page.get_pixmap(dpi=300)  # 提高DPI
                img = Image.frombytes("RGB", [page_image.width, page_image.height], page_image.samples)
                # 增强图像处理步骤
                img = ImageOps.grayscale(img)
                img = img.resize((img.width * 2, img.height * 2), Image.LANCZOS)
                enhancer = ImageEnhance.Contrast(img)
                img = enhancer.enhance(2)  # 调整对比度
                enhancer = ImageEnhance.Sharpness(img)
                img = enhancer.enhance(2)  # 调整锐度
                # 将图像转换为字符串
                text += pytesseract.image_to_string(img, lang='chi_sim+eng')

        doc.close()
        return text

    def urlToText(self, pdf_url):
        fileType = self.get_file_extension(pdf_url)
        file = self.download_file(pdf_url)
        text = self.fileToString(file,fileType)
        text = re.sub(r'\s+', ' ', text).strip()
        return text


if __name__ == '__main__':
    fileToText = FileToText()
    #文字型PDF
    #pdf_url = "https://xiaoxi-cdn.globeedu.com/2024/07/19/pdf/2024071915532394803013.pdf"  # 替换为你的PDF链接

    #图片型PDF
    #pdf_url = "https://xiaoxi-cdn.globeedu.com/2024/07/19/pdf/2024071915530687603013.pdf"  # 替换为你的PDF链接

    #DOC和DOCX
    #pdf_url = "http://xiaoxi-cdn.globeedu.com/2023/03/20/docx/2023032018443820303013.docx"
    #pdf_url = "D://123.doc"

    #PNG
    #pdf_url = "http://xiaoxi-cdn.globeedu.com/2023/04/17/png/2023041711055944503013.png"
    #JPG
    #pdf_url = "http://xiaoxi-cdn.globeedu.com/2023/04/17/jpg/2023041711083100403013.jpg"

    #XLSX
    pdf_url = "http://xiaoxi-cdn.globeedu.com/2023/03/20/xlsx/2023032018555339603013.xlsx"
    text = fileToText.urlToText(pdf_url)

    print(text)
